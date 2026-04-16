#!/usr/bin/env python3
"""
抓取兆豐-新莊對健亞(4130)從 2025/11 至今的完整買賣明細，存成 Excel。
"""

import requests
import urllib3
import re
import pandas as pd
from datetime import datetime

urllib3.disable_warnings()

STOCK_ID = "4130"
STOCK_NAME = "健亞"
BROKER_B = "0037003000300077"
BROKER_BHID = "7000"
BROKER_NAME = "兆豐-新莊"
START_DATE = "2025-11-01"

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}


def fetch_broker_data():
    """從 MoneyDJ 抓取券商分點歷史買賣資料"""
    today = datetime.now().strftime("%Y-%m-%d")
    url = (
        f"https://fubon-ebrokerdj.fbs.com.tw/z/zc/zco/zco0/zco0.djhtm"
        f"?A={STOCK_ID}&b={BROKER_B}&BHID={BROKER_BHID}"
        f"&C=1&D={START_DATE}&E={today}&ver=V3"
    )
    r = requests.get(url, headers=HEADERS, timeout=15, verify=False)
    r.encoding = "big5"

    pattern = (
        r'<TD class="t4n0">(20\d{2}/\d{2}/\d{2})</TD>\s*'
        r'<TD class="t3[nr]1">([\d,\-]+)</TD>\s*'
        r'<TD class="t3[nr]1">([\d,\-]+)</TD>\s*'
        r'<TD class="t3[nr]1">([\d,\-]+)</TD>\s*'
        r'<TD class="t3[nr]1">([\d,\-]+)</TD>'
    )
    matches = re.findall(pattern, r.text)

    data = {}
    for m in matches:
        buy = int(m[1].replace(",", ""))
        sell = int(m[2].replace(",", ""))
        data[m[0]] = {"買張": buy, "賣張": sell}

    print(f"券商資料：{len(data)} 筆")
    return data


def fetch_prices():
    """從 Yahoo Finance 取得歷史收盤價"""
    p1 = int(datetime(2025, 11, 1).timestamp())
    p2 = int(datetime.now().timestamp()) + 86400
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{STOCK_ID}.TWO"
        f"?period1={p1}&period2={p2}&interval=1d"
    )
    r = requests.get(url, headers=HEADERS, timeout=15)
    j = r.json()
    timestamps = j["chart"]["result"][0]["timestamp"]
    closes = j["chart"]["result"][0]["indicators"]["quote"][0]["close"]

    price_map = {}
    for ts, close in zip(timestamps, closes):
        dt = datetime.fromtimestamp(ts)
        date_key = dt.strftime("%Y/%m/%d")
        if close is not None:
            price_map[date_key] = round(close, 2)

    print(f"價格資料：{len(price_map)} 筆")
    return price_map


def main():
    broker_data = fetch_broker_data()
    price_map = fetch_prices()

    rows = []
    for date_str in sorted(broker_data.keys()):
        d = broker_data[date_str]
        price = price_map.get(date_str)
        buy_amt = round(d["買張"] * 1000 * price / 1000) if price and d["買張"] else 0
        sell_amt = round(d["賣張"] * 1000 * price / 1000) if price and d["賣張"] else 0
        rows.append({
            "日期": date_str,
            "買張": d["買張"],
            "賣張": d["賣張"],
            "買賣超(張)": d["買張"] - d["賣張"],
            "收盤價": price if price else "",
            "買金額(千元)": buy_amt,
            "賣金額(千元)": sell_amt,
            "買賣超金額(千元)": buy_amt - sell_amt,
        })

    df = pd.DataFrame(rows)

    # 存 Excel
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = "/Users/wuahe/Claude/追蹤健亞股票/兆豐新莊_健亞4130_交易明細.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="兆豐新莊-健亞4130", index=False)
        ws = writer.sheets["兆豐新莊-健亞4130"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

        col_widths = {"A": 14, "B": 12, "C": 12, "D": 14, "E": 10, "F": 16, "G": 16, "H": 18}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 合計列
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1, value="合計").font = Font(bold=True, size=12)
        for col, val in [
            (2, df["買張"].sum()),
            (3, df["賣張"].sum()),
            (4, df["買賣超(張)"].sum()),
            (6, df["買金額(千元)"].sum()),
            (7, df["賣金額(千元)"].sum()),
            (8, df["買賣超金額(千元)"].sum()),
        ]:
            ws.cell(row=last_row, column=col, value=val).font = Font(bold=True)
            ws.cell(row=last_row, column=col).number_format = "#,##0"

        for col in range(1, 9):
            ws.cell(row=last_row, column=col).border = thin_border
            ws.cell(row=last_row, column=col).alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

    print(f"\n已儲存：{output}")
    print(f"期間：{df['日期'].iloc[0]} ~ {df['日期'].iloc[-1]}，共 {len(df)} 筆")
    print(f"總買張：{df['買張'].sum():,}")
    print(f"總賣張：{df['賣張'].sum():,}")
    print(f"淨買超：{df['買賣超(張)'].sum():+,} 張")


if __name__ == "__main__":
    main()
